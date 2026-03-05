import asyncio
import os
from playwright.async_api import async_playwright
import openpyxl
import xml.etree.ElementTree as ET
from xml.dom import minidom
from datetime import datetime, date

# Config
DOWNLOAD_URL = "https://predajca.festool.sk/b5152248-3b3d-426f-b973-89b78b2022ca"
USERNAME = os.environ.get("FESTOOL_USERNAME", "")
PASSWORD = os.environ.get("FESTOOL_PASSWORD", "")
OUTPUT_FILE = "feed.xml"


async def screenshot(page, name):
    path = f"debug_{name}.png"
    await page.screenshot(path=path, full_page=True)
    print(f"   Screenshot: {path}")


async def download_excel():
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(
            accept_downloads=True,
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        )
        page = await context.new_page()

        # KROK 1: Najprv ideme na dealer portal - ten nas presmeruje na login s OAuth parametrami
        print("1. Otvarame dealer portal (presmeruje na login)...")
        await page.goto(DOWNLOAD_URL, wait_until="networkidle", timeout=60000)
        await page.wait_for_timeout(3000)
        await screenshot(page, "01_redirected_login")
        print(f"   URL: {page.url}")

        # COOKIE CONSENT - musime odkliknut pred loginom
        print("   Kontrolujem cookie consent banner...")
        try:
            await page.wait_for_timeout(2000)
            cookie_btn = await page.query_selector('button:has-text("Accept All")')
            if not cookie_btn:
                cookie_btn = await page.query_selector('button:has-text("Accept all")')
            if not cookie_btn:
                cookie_btn = await page.query_selector('button:has-text("Only accept necessary")')
            if cookie_btn and await cookie_btn.is_visible():
                await cookie_btn.click()
                print("   Cookie consent odkliknuty!")
                await page.wait_for_timeout(2000)
            else:
                print("   Cookie banner sa nezobrazil, pokracujeme...")
        except Exception as e:
            print(f"   Cookie handling: {e}")

        await screenshot(page, "01b_after_cookies")

        # Overime ze sme na login stranke
        if "login.festool.com" not in page.url:
            print("   Uz sme prihlaseni, pokracujeme...")
        else:
            # KROK 2: Vyplnime udaje na login stranke (uz s korektnym ReturnUrl)
            print("2. Zadavame prihlasovacie udaje...")
            await page.wait_for_timeout(2000)

            # Skusame rozne selektory pre email
            email_filled = False
            for sel in ['input[name="Input.Email"]', 'input[name="Email"]', 'input[type="email"]', '#Email', '#Input_Email', 'input[id*="mail"]']:
                try:
                    el = await page.query_selector(sel)
                    if el and await el.is_visible():
                        await el.fill(USERNAME)
                        print(f"   Email vyplneny cez: {sel}")
                        email_filled = True
                        break
                except:
                    continue

            if not email_filled:
                # Skusame vsetky textove inputy
                inputs = await page.query_selector_all('input[type="text"], input[type="email"], input:not([type])')
                for inp in inputs:
                    try:
                        if await inp.is_visible():
                            placeholder = await inp.get_attribute("placeholder") or ""
                            name_attr = await inp.get_attribute("name") or ""
                            if any(w in (placeholder + name_attr).lower() for w in ['email', 'mail', 'user', 'login', 'meno']):
                                await inp.fill(USERNAME)
                                print(f"   Email vyplneny cez placeholder/name hladanie")
                                email_filled = True
                                break
                    except:
                        continue

            if not email_filled:
                # Posledna moznost - prvy viditelny textovy input
                inputs = await page.query_selector_all('input[type="text"], input[type="email"], input:not([type="hidden"]):not([type="password"]):not([type="submit"]):not([type="checkbox"])')
                for inp in inputs:
                    try:
                        if await inp.is_visible():
                            await inp.fill(USERNAME)
                            print(f"   Email vyplneny do prveho viditelneho inputu")
                            email_filled = True
                            break
                    except:
                        continue

            # Heslo
            password_filled = False
            for sel in ['input[name="Input.Password"]', 'input[name="Password"]', 'input[type="password"]', '#Password', '#Input_Password']:
                try:
                    el = await page.query_selector(sel)
                    if el and await el.is_visible():
                        await el.fill(PASSWORD)
                        print(f"   Heslo vyplnene cez: {sel}")
                        password_filled = True
                        break
                except:
                    continue

            if not email_filled or not password_filled:
                print(f"   VAROVANIE: email_filled={email_filled}, password_filled={password_filled}")
                # Debug - vypiseme vsetky inputy na stranke
                all_inputs = await page.query_selector_all('input')
                for inp in all_inputs:
                    try:
                        inp_type = await inp.get_attribute("type") or "none"
                        inp_name = await inp.get_attribute("name") or "none"
                        inp_id = await inp.get_attribute("id") or "none"
                        visible = await inp.is_visible()
                        print(f"   INPUT: type={inp_type}, name={inp_name}, id={inp_id}, visible={visible}")
                    except:
                        pass

            await screenshot(page, "02_filled_form")

            # KROK 3: Prihlasenie
            print("3. Prihlasujeme sa...")
            clicked = False
            for sel in ['button[type="submit"]', 'input[type="submit"]', '.btn-primary', 'button.btn', '#login-button']:
                try:
                    el = await page.query_selector(sel)
                    if el and await el.is_visible():
                        await el.click()
                        print(f"   Kliknute na: {sel}")
                        clicked = True
                        break
                except:
                    continue

            if not clicked:
                # Skusame Enter
                await page.keyboard.press("Enter")
                print("   Odoslane cez Enter")

            await page.wait_for_timeout(5000)
            await page.wait_for_load_state("networkidle", timeout=60000)
            await screenshot(page, "03_after_login")
            print(f"   URL po prihlaseni: {page.url}")

            # Ak sme stale na login, skusime este raz navigovat na DOWNLOAD_URL
            if "login.festool.com" in page.url:
                print("   Stale na login stranke, skusame znova navigovat...")
                await page.goto(DOWNLOAD_URL, wait_until="networkidle", timeout=60000)
                await page.wait_for_timeout(5000)
                await screenshot(page, "03b_retry_navigate")
                print(f"   URL po retry: {page.url}")

                if "login.festool.com" in page.url:
                    print("   CHYBA: Prihlasenie zlyhalo!")
                    # Vypiseme obsah stranky pre debug
                    page_content = await page.content()
                    with open("debug_page.html", "w", encoding="utf-8") as f:
                        f.write(page_content)
                    print("   HTML stranky ulozene do debug_page.html")
                    await browser.close()
                    return None

        # KROK 4: Sme na dealer portali - hladame "Excel export všetko"
        print("4. Sme na dealer portali, hladame export tlacidlo...")
        await page.wait_for_timeout(3000)

        # Cookie consent aj na dealer portali
        try:
            cookie_btn = await page.query_selector('button:has-text("Accept All")')
            if not cookie_btn:
                cookie_btn = await page.query_selector('button:has-text("Accept all")')
            if not cookie_btn:
                cookie_btn = await page.query_selector('button:has-text("Only accept necessary")')
            if cookie_btn and await cookie_btn.is_visible():
                await cookie_btn.click()
                print("   Cookie consent na portali odkliknuty!")
                await page.wait_for_timeout(2000)
        except:
            pass

        await screenshot(page, "04_dealer_portal")
        print(f"   URL: {page.url}")

        # Najprv skusame presny text tlacidla z obrazovky
        download_path = None

        # Hladame tlacidlo "Excel export všetko"
        export_selectors = [
            'button:has-text("Excel export všetko")',
            'a:has-text("Excel export všetko")',
            'button:has-text("Excel export v")',
            'a:has-text("Excel export v")',
            'button:has-text("export všetko")',
            'a:has-text("export všetko")',
            'button:has-text("Excel export")',
            'a:has-text("Excel export")',
        ]

        for sel in export_selectors:
            try:
                el = await page.query_selector(sel)
                if el and await el.is_visible():
                    print(f"   Nasiel som tlacidlo: {sel}")
                    async with page.expect_download(timeout=120000) as download_info:
                        await el.click()
                    download = await download_info.value
                    download_path = f"downloads/{download.suggested_filename}"
                    os.makedirs("downloads", exist_ok=True)
                    await download.save_as(download_path)
                    print(f"   Subor stiahnuty: {download_path}")
                    break
            except Exception as e:
                print(f"   Selektor {sel} - chyba: {e}")
                continue

        # Ak to neslo cez text, skusame vsetky buttony a linky
        if not download_path:
            print("   Prehladavame vsetky elementy...")
            elements = await page.query_selector_all('button, a, [role="button"], input[type="button"], input[type="submit"]')
            for el in elements:
                try:
                    text = (await el.inner_text()).strip().lower()
                    if any(w in text for w in ['export všetko', 'export vsetko', 'excel export', 'export all']):
                        print(f"   Nasiel som element s textom: {text}")
                        async with page.expect_download(timeout=120000) as download_info:
                            await el.click()
                        download = await download_info.value
                        download_path = f"downloads/{download.suggested_filename}"
                        os.makedirs("downloads", exist_ok=True)
                        await download.save_as(download_path)
                        print(f"   Subor stiahnuty cez text match: {download_path}")
                        break
                except:
                    continue

        # Link s .xlsx alebo download atributom
        if not download_path:
            links = await page.query_selector_all('a[href*=".xlsx"], a[href*="download"], a[href*="export"], a[download]')
            for link in links:
                try:
                    async with page.expect_download(timeout=60000) as download_info:
                        await link.click()
                    download = await download_info.value
                    download_path = f"downloads/{download.suggested_filename}"
                    os.makedirs("downloads", exist_ok=True)
                    await download.save_as(download_path)
                    print(f"   Subor stiahnuty cez href: {download_path}")
                    break
                except Exception as e:
                    print(f"   Link nefungoval: {e}")
                    continue

        if not download_path:
            await screenshot(page, "05_no_download")
            print("   CHYBA: Nepodarilo sa najst/stiahnut subor!")
            page_content = await page.content()
            with open("debug_page.html", "w", encoding="utf-8") as f:
                f.write(page_content)
            print("   HTML stranky ulozene do debug_page.html")
            # Vypiseme vsetky viditelne buttony/linky
            elements = await page.query_selector_all('button, a')
            for el in elements:
                try:
                    text = (await el.inner_text()).strip()
                    if text:
                        href = await el.get_attribute("href") or ""
                        print(f"   ELEMENT: '{text}' href='{href}'")
                except:
                    pass
            await browser.close()
            return None

        await browser.close()
        return download_path


def generate_feed(excel_path):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb[wb.sheetnames[0]]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    today = date.today()
    print(f"Dnesny datum: {today}")

    root = ET.Element("CHANNEL")
    root.set("xmlns", "http://www.mergado.com/ns/1.10")

    ET.SubElement(root, "LINK").text = "https://github.com/Roman-Hudak/aspotrmarre"
    ET.SubElement(root, "GENERATOR").text = "custom_feed_generator_1.0"

    count = 0
    in_stock = 0
    out_stock = 0

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        data = dict(zip(headers, row))
        if not data.get('Obj. číslo') or not data.get('Opis'):
            continue

        item = ET.SubElement(root, "ITEM")
        ET.SubElement(item, "ITEM_ID").text = str(data.get('Obj. číslo', ''))
        ET.SubElement(item, "NAME_EXACT").text = str(data.get('Opis', ''))

        if data.get('Cena EUR') is not None:
            ET.SubElement(item, "PRICE_VAT").text = str(data['Cena EUR'])
        if data.get('Netto NC EUR') is not None:
            ET.SubElement(item, "COST").text = str(data['Netto NC EUR'])

        ET.SubElement(item, "CURRENCY").text = "EUR"

        if data.get('EAN'):
            ET.SubElement(item, "EAN").text = str(data['EAN'])
        if data.get('Typ'):
            ET.SubElement(item, "CATEGORY").text = str(data['Typ'])
        if data.get('Hierarchia produktov'):
            ET.SubElement(item, "CATEGORYTEXT").text = str(data['Hierarchia produktov'])

        for label, key in [('Výška', 'Výška'), ('Šírka', 'Šírka'), ('Dĺžka', 'Dĺžka'), ('Hmotnosť', 'Hmotnosť')]:
            if data.get(key):
                param = ET.SubElement(item, "PARAM")
                ET.SubElement(param, "n").text = label
                ET.SubElement(param, "VALUE").text = str(data[key])

        if data.get('CoO'):
            ET.SubElement(item, "COUNTRY_OF_ORIGIN").text = str(data['CoO'])

        # AVAILABILITY: in stock iba ak datum dodania <= dnes
        delivery_date = data.get('Dátum dodania')
        if delivery_date and hasattr(delivery_date, 'date'):
            delivery = delivery_date.date()
        elif delivery_date and hasattr(delivery_date, 'strftime'):
            delivery = delivery_date
        else:
            delivery = None

        if delivery and delivery <= today:
            ET.SubElement(item, "AVAILABILITY").text = "in stock"
            in_stock += 1
        else:
            ET.SubElement(item, "AVAILABILITY").text = "out of stock"
            out_stock += 1

        ET.SubElement(item, "CONDITION").text = "new"

        if data.get('Partnerská zľava') is not None:
            param = ET.SubElement(item, "PARAM")
            ET.SubElement(param, "n").text = "Partnerská zľava"
            ET.SubElement(param, "VALUE").text = str(data['Partnerská zľava'])

        if delivery:
            ET.SubElement(item, "DELIVERY_DATE").text = delivery.strftime('%Y-%m-%d')

        if data.get('Toolpoints'):
            param = ET.SubElement(item, "PARAM")
            ET.SubElement(param, "n").text = "Toolpoints"
            ET.SubElement(param, "VALUE").text = str(data['Toolpoints'])

        count += 1

    xml_str = minidom.parseString(ET.tostring(root, encoding='unicode')).toprettyxml(indent="  ")
    lines = xml_str.split(chr(10))
    xml_str = '<?xml version="1.0" encoding="utf-8"?>\n' + chr(10).join(lines[1:])

    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        f.write(xml_str)

    print(f"Feed vygenerovany: {count} produktov")
    print(f"  IN STOCK:     {in_stock}")
    print(f"  OUT OF STOCK: {out_stock}")
    return count


async def main():
    os.makedirs("downloads", exist_ok=True)

    print("=" * 50)
    print("FESTOOL XML FEED GENERATOR")
    print("=" * 50)

    excel_path = await download_excel()

    if excel_path:
        count = generate_feed(excel_path)
        if count > 0:
            print(f"\nUSPECH! Feed s {count} produktmi bol vygenerovany.")
        else:
            print("\nCHYBA: Ziadne produkty neboli najdene v subore.")
            exit(1)
    else:
        print("\nCHYBA: Nepodarilo sa stiahnut Excel subor.")
        exit(1)

if __name__ == "__main__":
    asyncio.run(main())
